from pathlib import Path

import geopandas as gpd
import numpy as np
import polars as pl
import pytest
from openpyxl import Workbook, load_workbook
from shapely.geometry import box

from cadence.reference_data.year1_damage import (
    GUST_COLUMNS,
    _join_asset_results,
    attach_climate_zones,
    attach_nearest_wind_grid,
    build_asset_scoped_damage,
    calculate_asset_damage,
    read_asset_workbook,
)
from cadence.vulnerability.expected_damage import (
    MPS_TO_MPH,
    integrate_damage_matrix,
    interpolate_return_period_damage,
)


def _write_assets(path: Path, rows: list) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(
        [
            "asset_id",
            "latitude",
            "longitude",
            "current_roof_type",
            "roof_age",
            "Terrian",
            "Terrian_numeric",
        ]
    )
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def test_reads_legacy_terrain_headers_and_caps_age(tmp_path: Path) -> None:
    path = tmp_path / "assets.xlsx"
    _write_assets(path, [["A-1", 32.0, -80.0, "Metal", 42, "Suburban", 3]])

    assets = read_asset_workbook(path)

    assert assets.row(0, named=True) == {
        "asset_id": "A-1",
        "latitude": 32.0,
        "longitude": -80.0,
        "current_roof_type": "metal",
        "input_roof_age": 42,
        "lookup_roof_age": 30,
        "age_capped": True,
        "terrain_id": 3,
        "terrain_label": "Suburban",
    }


def test_rejects_disagreeing_terrain_label_and_code(tmp_path: Path) -> None:
    path = tmp_path / "assets.xlsx"
    _write_assets(path, [["A-1", 32.0, -80.0, "Asphalt", 5, "Open", 3]])

    with pytest.raises(ValueError, match="do not agree"):
        read_asset_workbook(path)


@pytest.mark.parametrize(
    ("label", "terrain_id"),
    [
        ("Open", 1),
        ("light suburban", 2),
        ("SUBURBAN", 3),
        ("Light Trees", 4),
        ("trees", 5),
    ],
)
def test_reads_all_canonical_terrain_values(
    tmp_path: Path, label: str, terrain_id: int
) -> None:
    path = tmp_path / "assets.xlsx"
    _write_assets(path, [["A-1", 32.0, -80.0, "Asphalt", 5, label, terrain_id]])

    asset = read_asset_workbook(path).row(0, named=True)

    assert asset["terrain_id"] == terrain_id


@pytest.mark.parametrize(
    "row",
    [
        ["A-1", float("nan"), -80.0, "Asphalt", 5, "Open", 1],
        ["A-1", 32.0, -80.0, "Asphalt", 5, "Industrial", None],
        ["A-1", 32.0, -80.0, "Asphalt", 5, None, 6],
    ],
)
def test_rejects_nonfinite_coordinates_and_invalid_terrain(
    tmp_path: Path, row: list
) -> None:
    path = tmp_path / "assets.xlsx"
    _write_assets(path, [row])

    with pytest.raises(ValueError):
        read_asset_workbook(path)


def test_reads_corrected_repository_workbook() -> None:
    path = Path("Data/User_Inputs/asset_inventory_test_1.xlsx")

    assets = read_asset_workbook(path)

    assert assets.height == 2
    assert assets["terrain_id"].to_list() == [1, 1]
    assert assets["terrain_label"].to_list() == ["Open", "Open"]
    assert assets["input_roof_age"].to_list() == [15, 5]


def test_attaches_nearest_wind_grid_once_per_location(tmp_path: Path) -> None:
    wind_path = tmp_path / "wind.csv"
    wind_rows = {
        "lat_idx": [1, 2],
        "lon_idx": [10, 20],
        "latitude": [32.0, 40.0],
        "longitude": [-80.0, -100.0],
        "converged": [1, 1],
    }
    wind_rows.update({column: [20.0, 30.0] for column in GUST_COLUMNS})
    pl.DataFrame(wind_rows).write_csv(wind_path)
    assets = pl.DataFrame(
        {
            "asset_id": ["A-1", "A-2"],
            "latitude": [32.01, 32.01],
            "longitude": [-80.01, -80.01],
        }
    )

    attached = attach_nearest_wind_grid(assets, wind_path)

    assert attached.height == 2
    assert attached["wind_grid_id"].to_list() == [65_546, 65_546]
    assert attached["rp_10_3sec_gust"].to_list() == pytest.approx(
        [20.0 * MPS_TO_MPH, 20.0 * MPS_TO_MPH]
    )
    assert max(attached["wind_grid_distance_km"]) < 2.0


def test_attaches_distinct_locations_to_distinct_wind_cells(tmp_path: Path) -> None:
    wind_path = tmp_path / "wind.csv"
    wind_rows = {
        "lat_idx": [1, 2],
        "lon_idx": [10, 20],
        "latitude": [32.0, 40.0],
        "longitude": [-80.0, -100.0],
        "converged": [1, 1],
    }
    wind_rows.update({column: [20.0, 30.0] for column in GUST_COLUMNS})
    pl.DataFrame(wind_rows).write_csv(wind_path)
    assets = pl.DataFrame(
        {
            "asset_id": ["A-1", "A-2"],
            "latitude": [32.01, 40.01],
            "longitude": [-80.01, -100.01],
        }
    )

    attached = attach_nearest_wind_grid(assets, wind_path)

    assert attached["wind_grid_id"].to_list() == [65_546, 131_092]
    assert attached["rp_10_3sec_gust"].to_list() == pytest.approx(
        [20.0 * MPS_TO_MPH, 30.0 * MPS_TO_MPH]
    )


def test_rejects_nonfinite_wind_values(tmp_path: Path) -> None:
    wind_path = tmp_path / "wind.csv"
    wind_rows = {
        "lat_idx": [1],
        "lon_idx": [10],
        "latitude": [32.0],
        "longitude": [-80.0],
        "converged": [1],
    }
    wind_rows.update({column: [20.0] for column in GUST_COLUMNS})
    wind_rows[GUST_COLUMNS[0]] = [float("nan")]
    pl.DataFrame(wind_rows).write_csv(wind_path)
    assets = pl.DataFrame({"asset_id": ["A-1"], "latitude": [32.0], "longitude": [-80.0]})

    with pytest.raises(ValueError, match="non-finite"):
        attach_nearest_wind_grid(assets, wind_path)


def test_assigns_and_rejects_climate_zones(monkeypatch: pytest.MonkeyPatch) -> None:
    zones = gpd.GeoDataFrame(
        {"IECC21": [3], "Moisture21": ["A"]},
        geometry=[box(-81.0, 31.0, -79.0, 33.0)],
        crs="EPSG:4326",
    )
    monkeypatch.setattr("cadence.reference_data.year1_damage.gpd.read_file", lambda _: zones)
    resolved = pl.DataFrame(
        {"asset_id": ["A-1"], "latitude": [32.0], "longitude": [-80.0]}
    )

    assert attach_climate_zones(resolved, Path("unused"))["climate_zone"].item() == "3A"

    unresolved = pl.DataFrame(
        {"asset_id": ["A-2"], "latitude": [40.0], "longitude": [-100.0]}
    )
    with pytest.raises(ValueError, match="A-2"):
        attach_climate_zones(unresolved, Path("unused"))


def test_calculates_three_materials_once_for_shared_key(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    rows = {
        "asset_id": ["A-1", "A-2"],
        "wind_grid_id": [1, 1],
        "climate_zone": ["3A", "3A"],
        "lookup_roof_age": [5, 5],
        "terrain_id": [3, 3],
    }
    rows.update({column: [30.0, 30.0] for column in GUST_COLUMNS})
    assets = pl.DataFrame(rows)
    calls = []

    def fake_curve(
        fragility_root: Path,
        climate_zone: str,
        roof_age: int,
        official_material_id: str,
        terrain_id: int,
    ) -> pl.DataFrame:
        calls.append((climate_zone, roof_age, official_material_id, terrain_id))
        return pl.DataFrame(
            {
                "wind_speed_mph": [50.0, 100.0],
                "building_loss_ratio": [0.0, 0.5],
                "source_curve_count": [10, 10],
            }
        )

    monkeypatch.setattr(
        "cadence.reference_data.year1_damage.load_terrain_averaged_curve",
        fake_curve,
    )

    lookup, results = calculate_asset_damage(assets, Path("unused"))

    assert len(calls) == 3
    assert lookup.height == 3
    assert results.height == 6
    assert results.group_by("asset_id").len()["len"].to_list() == [3, 3]


def test_cached_lookup_rejects_missing_key() -> None:
    assets = pl.DataFrame(
        {
            "asset_id": ["A-1"],
            "wind_grid_id": [1],
            "climate_zone": ["3A"],
            "lookup_roof_age": [5],
            "terrain_id": [3],
        }
    )
    lookup = pl.DataFrame(
        {
            "wind_grid_id": [2, 2, 2],
            "climate_zone": ["3A"] * 3,
            "lookup_roof_age": [5] * 3,
            "terrain_id": [3] * 3,
            "official_material_id": [
                "OFFICIAL_ASPHALT",
                "OFFICIAL_METAL",
                "OFFICIAL_TILE",
            ],
        }
    )

    with pytest.raises(ValueError, match="missing 1 required"):
        _join_asset_results(assets, lookup)


def test_cached_lookup_rejects_duplicate_or_incomplete_materials() -> None:
    assets = pl.DataFrame(
        {
            "asset_id": ["A-1"],
            "wind_grid_id": [1],
            "climate_zone": ["3A"],
            "lookup_roof_age": [5],
            "terrain_id": [3],
        }
    )
    lookup = pl.DataFrame(
        {
            "wind_grid_id": [1, 1, 1],
            "climate_zone": ["3A"] * 3,
            "lookup_roof_age": [5] * 3,
            "terrain_id": [3] * 3,
            "official_material_id": [
                "OFFICIAL_ASPHALT",
                "OFFICIAL_ASPHALT",
                "OFFICIAL_TILE",
            ],
        }
    )

    with pytest.raises(ValueError, match="exactly Asphalt, Metal, and Tile"):
        _join_asset_results(assets, lookup)


def test_integrates_hand_calculated_aep_curve() -> None:
    damages = np.asarray([[0.2, 0.3, 0.4, 0.5, 0.6, 0.7]])
    expected = (
        0.9 * 0.2 / 2
        + 0.06 * (0.2 + 0.3) / 2
        + 0.02 * (0.3 + 0.4) / 2
        + 0.01 * (0.4 + 0.5) / 2
        + 0.005 * (0.5 + 0.6) / 2
        + 0.003 * (0.6 + 0.7) / 2
        + 0.002 * (0.7 + 1.0) / 2
    )

    assert integrate_damage_matrix(damages)[0] == pytest.approx(expected)


def test_rejects_nan_damage_and_gust_values() -> None:
    with pytest.raises(ValueError, match="finite"):
        integrate_damage_matrix(np.asarray([[0.0, 0.1, np.nan, 0.3, 0.4, 0.5]]))
    with pytest.raises(ValueError, match="finite"):
        interpolate_return_period_damage(
            [1.0, 2.0, np.nan, 4.0, 5.0, 6.0],
            np.asarray([50.0, 100.0]),
            np.asarray([0.0, 1.0]),
        )


def test_exact_curve_endpoints_are_not_counted_as_clamps() -> None:
    gust_mph = [50.0, 60.0, 70.0, 80.0, 90.0, 100.0]

    damages, lower_clamps, upper_clamps = interpolate_return_period_damage(
        gust_mph,
        np.asarray([50.0, 100.0]),
        np.asarray([0.0, 1.0]),
    )

    assert lower_clamps == 0
    assert upper_clamps == 0
    assert damages[[0, -1]].tolist() == pytest.approx([0.0, 1.0])


def test_different_portfolios_reuse_calculation_cache_without_sharing_assets(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    current_assets = ["A-1", "A-2"]

    def prepared_assets() -> pl.DataFrame:
        rows = {
            "asset_id": current_assets,
            "latitude": [32.0, 32.0],
            "longitude": [-80.0, -80.0],
            "current_roof_type": ["asphalt", "metal"],
            "input_roof_age": [5, 5],
            "lookup_roof_age": [5, 5],
            "age_capped": [False, False],
            "terrain_id": [3, 3],
            "terrain_label": ["Suburban", "Suburban"],
            "wind_grid_id": [1, 1],
            "climate_zone": ["3A", "3A"],
        }
        rows.update({column: [30.0, 30.0] for column in GUST_COLUMNS})
        return pl.DataFrame(rows)

    monkeypatch.setattr(
        "cadence.reference_data.year1_damage.read_asset_workbook",
        lambda *_: prepared_assets(),
    )
    monkeypatch.setattr(
        "cadence.reference_data.year1_damage.attach_nearest_wind_grid",
        lambda assets, _: assets,
    )
    monkeypatch.setattr(
        "cadence.reference_data.year1_damage.attach_climate_zones",
        lambda assets, _: assets,
    )
    monkeypatch.setattr(
        "cadence.reference_data.year1_damage._fragility_identity",
        lambda *_: "fragility-test-identity",
    )

    def fake_calculation(assets: pl.DataFrame, _: Path) -> tuple:
        lookup = pl.DataFrame(
            {
                "wind_grid_id": [1, 1, 1],
                "climate_zone": ["3A"] * 3,
                "lookup_roof_age": [5] * 3,
                "terrain_id": [3] * 3,
                "official_material_id": [
                    "OFFICIAL_ASPHALT",
                    "OFFICIAL_METAL",
                    "OFFICIAL_TILE",
                ],
                "year_1_expected_damage_ratio": [0.1, 0.08, 0.09],
            }
        )
        return lookup, _join_asset_results(assets, lookup)

    monkeypatch.setattr(
        "cadence.reference_data.year1_damage.calculate_asset_damage", fake_calculation
    )
    arguments = (
        Path("assets.xlsx"),
        Path("wind.csv"),
        Path("zones.shp"),
        Path("fragility"),
        tmp_path,
    )

    first = build_asset_scoped_damage(*arguments)
    current_assets[:] = ["B-1", "B-2"]
    second = build_asset_scoped_damage(*arguments)
    second_results = pl.read_parquet(second["asset_results_path"])

    assert not first["calculation_cache_hit"]
    assert second["calculation_cache_hit"]
    assert sorted(second_results["asset_id"].unique()) == ["B-1", "B-2"]
    assert len(list(tmp_path.rglob("unique_location_damage_lookup.parquet"))) == 1


@pytest.mark.integration
def test_repository_data_end_to_end_with_corrected_terrain_workbook(
    tmp_path: Path,
) -> None:
    source = Path("Data/User_Inputs/asset_inventory_test_1.xlsx")
    workbook_path = tmp_path / "asset_inventory_onshore.xlsx"
    workbook = load_workbook(source)
    sheet = workbook["Sheet1"]
    for row, (latitude, longitude) in enumerate(
        ((32.7765, -79.9311), (33.8361, -80.8987)), start=2
    ):
        sheet.cell(row, 2, latitude)
        sheet.cell(row, 3, longitude)
    workbook.save(workbook_path)

    manifest = build_asset_scoped_damage(
        workbook_path,
        Path("Data/Wind_Return_Periods/gev_return_periods.csv"),
        Path("Data/Climate_Zones/ClimateZones.shp"),
        Path("Data/Fragility_Curves/iecc2021"),
        tmp_path / "results",
    )
    lookup = pl.read_parquet(manifest["lookup_path"])
    results = pl.read_parquet(manifest["asset_results_path"])

    assert manifest["asset_count"] == 2
    assert manifest["unique_calculation_key_count"] == 2
    assert lookup.height == 6
    assert results.height == 6
    assert set(results["official_material_id"]) == {
        "OFFICIAL_ASPHALT",
        "OFFICIAL_METAL",
        "OFFICIAL_TILE",
    }
    assert results["terrain_id"].unique().to_list() == [1]
    assert results["terrain_label"].unique().to_list() == ["Open"]
    assert results["year_1_expected_damage_ratio"].is_finite().all()