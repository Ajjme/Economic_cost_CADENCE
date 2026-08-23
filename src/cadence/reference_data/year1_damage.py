"""Calculate baseline Year 1 roof damage only for user-supplied asset locations."""

import argparse
import hashlib
import json
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import geopandas as gpd
import numpy as np
import polars as pl
from openpyxl import load_workbook
from scipy.spatial import cKDTree

from cadence.reference_data.fragility import (
    FRAGILITY_MAPPING_VERSION,
    MATERIAL_PROXY_IDS,
    TERRAIN_IDS_BY_LABEL,
    TERRAIN_LABELS,
    load_terrain_averaged_curve,
)
from cadence.vulnerability.expected_damage import (
    MPS_TO_MPH,
    RETURN_PERIODS,
    integrate_damage_matrix,
    interpolate_return_period_damage,
)

EXPECTED_DAMAGE_VERSION = "asset-scoped-baseline-year1-v0.2.1"
GUST_COLUMNS = tuple(f"rp_{period}_3sec_gust" for period in RETURN_PERIODS)
ASSET_REQUIRED_COLUMNS = ("asset_id", "latitude", "longitude", "roof_age")
TERRAIN_LABEL_ALIASES = ("terrain", "Terrain", "Terrian")
TERRAIN_ID_ALIASES = (
    "terrain_id",
    "terrain_numeric",
    "Terrain_numeric",
    "Terrian_numeric",
)
VALID_CLIMATE_ZONES = (
    "1A", "2A", "2B", "3A", "3B", "3C", "4A", "4B", "4C",
    "5A", "5B", "5C", "6A", "6B", "7", "8",
)


def read_asset_workbook(path: Path, sheet_name: str = "Sheet1") -> pl.DataFrame:
    """Read and validate asset coordinates, age, and HAZUS terrain from Excel."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"asset sheet {sheet_name!r} not found; choices: {workbook.sheetnames}")
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration as error:
        raise ValueError("asset workbook is empty") from error
    headers = [str(value).strip() if value is not None else "" for value in raw_headers]
    records = [
        dict(zip(headers, row))
        for row in rows
        if any(value is not None and str(value).strip() for value in row)
    ]
    if not records:
        raise ValueError("asset workbook contains no asset rows")
    missing = [column for column in ASSET_REQUIRED_COLUMNS if column not in headers]
    if missing:
        raise ValueError(f"asset workbook is missing required columns: {missing}")

    terrain_label_column = _first_present(headers, TERRAIN_LABEL_ALIASES)
    terrain_id_column = _first_present(headers, TERRAIN_ID_ALIASES)
    if terrain_label_column is None and terrain_id_column is None:
        raise ValueError(
            "asset workbook requires terrain or terrain_id; accepted legacy headers include "
            "Terrian and Terrian_numeric"
        )

    normalized = []
    seen_asset_ids = set()
    for excel_row, record in enumerate(records, start=2):
        asset_id = str(record["asset_id"]).strip()
        if not asset_id or asset_id in seen_asset_ids:
            raise ValueError(f"row {excel_row}: asset_id must be nonblank and unique")
        seen_asset_ids.add(asset_id)
        latitude = _number(record["latitude"], excel_row, "latitude")
        longitude = _number(record["longitude"], excel_row, "longitude")
        roof_age_value = _number(record["roof_age"], excel_row, "roof_age")
        if not -90.0 <= latitude <= 90.0 or not -180.0 <= longitude <= 180.0:
            raise ValueError(f"row {excel_row}: coordinates are outside WGS84 bounds")
        if not roof_age_value.is_integer() or roof_age_value < 1:
            raise ValueError(f"row {excel_row}: roof_age must be a positive integer")
        input_age = int(roof_age_value)
        terrain_id = _resolve_terrain(record, terrain_label_column, terrain_id_column, excel_row)
        normalized.append(
            {
                "asset_id": asset_id,
                "latitude": latitude,
                "longitude": longitude,
                "current_roof_type": (
                    str(record.get("current_roof_type")).strip().casefold()
                    if record.get("current_roof_type") is not None
                    else None
                ),
                "input_roof_age": input_age,
                "lookup_roof_age": min(input_age, 30),
                "age_capped": input_age > 30,
                "terrain_id": terrain_id,
                "terrain_label": TERRAIN_LABELS[terrain_id],
            }
        )
    return pl.DataFrame(normalized)


def attach_nearest_wind_grid(assets: pl.DataFrame, wind_csv_path: Path) -> pl.DataFrame:
    """Attach the nearest CONUS404 center and six gust return levels to each asset."""
    wind = pl.read_csv(
        wind_csv_path,
        columns=["lat_idx", "lon_idx", "latitude", "longitude", *GUST_COLUMNS, "converged"],
    )
    wind = wind.with_columns(
        *(pl.col(column).mul(MPS_TO_MPH).alias(column) for column in GUST_COLUMNS)
    )
    finite_wind_columns = ["latitude", "longitude", *GUST_COLUMNS]
    invalid_wind = wind.filter(
        pl.any_horizontal(~pl.col(column).is_finite() for column in finite_wind_columns)
    )
    if invalid_wind.height:
        raise ValueError("wind grid contains non-finite coordinates or gust return levels")
    wind_xyz = _unit_sphere_xyz(wind["latitude"].to_numpy(), wind["longitude"].to_numpy())
    unique_locations = assets.select("latitude", "longitude").unique(maintain_order=True)
    asset_xyz = _unit_sphere_xyz(
        unique_locations["latitude"].to_numpy(), unique_locations["longitude"].to_numpy()
    )
    chord_distance, indices = cKDTree(wind_xyz).query(asset_xyz, k=1)
    earth_radius_km = 6371.0088
    distance_km = 2.0 * earth_radius_km * np.arcsin(np.clip(chord_distance / 2.0, 0.0, 1.0))
    nearest = wind[indices].with_columns(
        pl.Series("asset_latitude", unique_locations["latitude"]),
        pl.Series("asset_longitude", unique_locations["longitude"]),
        pl.Series("wind_grid_distance_km", distance_km),
        (
            pl.col("lat_idx").cast(pl.UInt64) * 65_536
            + pl.col("lon_idx").cast(pl.UInt64)
        ).alias("wind_grid_id"),
    ).rename({"latitude": "wind_grid_latitude", "longitude": "wind_grid_longitude"})
    attached = assets.join(
        nearest,
        left_on=["latitude", "longitude"],
        right_on=["asset_latitude", "asset_longitude"],
        how="left",
        validate="m:1",
    )
    required_columns = ["wind_grid_id", "wind_grid_latitude", "wind_grid_longitude", *GUST_COLUMNS]
    if attached.select(
        pl.any_horizontal(pl.col(column).is_null() for column in required_columns).any()
    ).item():
        raise ValueError("one or more assets could not be attached to the wind grid")
    return attached


def attach_climate_zones(assets: pl.DataFrame, climate_zone_path: Path) -> pl.DataFrame:
    """Assign each asset point to one valid IECC 2021 climate zone."""
    zones = gpd.read_file(climate_zone_path)[["IECC21", "Moisture21", "geometry"]]
    zones = zones.to_crs("EPSG:4326")
    numeric_zone = zones["IECC21"].astype("Int64")
    zone_number = numeric_zone.astype("string")
    moisture = zones["Moisture21"].astype("string").str.strip().str.upper()
    zones["climate_zone"] = zone_number.where(
        ~numeric_zone.between(1, 6), zone_number + moisture
    )
    zones = zones[zones["climate_zone"].isin(VALID_CLIMATE_ZONES)][
        ["climate_zone", "geometry"]
    ]
    asset_pandas = assets.to_pandas()
    points = gpd.GeoDataFrame(
        asset_pandas,
        geometry=gpd.points_from_xy(asset_pandas["longitude"], asset_pandas["latitude"]),
        crs="EPSG:4326",
    )
    joined = gpd.sjoin(points, zones, how="left", predicate="within")
    if len(joined) != len(points):
        raise ValueError("an asset point matched more than one climate-zone polygon")
    result = pl.from_pandas(joined.drop(columns=["geometry", "index_right"]))
    unresolved = result.filter(pl.col("climate_zone").is_null())
    if unresolved.height:
        raise ValueError(
            "climate zone could not be resolved for asset_ids: "
            + ", ".join(unresolved["asset_id"].to_list())
        )
    return result


def calculate_asset_damage(
    assets_with_hazard: pl.DataFrame,
    fragility_root: Path,
) -> Tuple[pl.DataFrame, pl.DataFrame]:
    """Calculate three materials once per unique location/zone/age/terrain key."""
    key_columns = ["wind_grid_id", "climate_zone", "lookup_roof_age", "terrain_id"]
    source_columns = [*key_columns, *GUST_COLUMNS]
    unique_keys = assets_with_hazard.select(source_columns).unique()
    calculation_rows: List[dict] = []
    for key in unique_keys.iter_rows(named=True):
        for official_material_id, proxy_id in MATERIAL_PROXY_IDS.items():
            curve = load_terrain_averaged_curve(
                fragility_root,
                key["climate_zone"],
                key["lookup_roof_age"],
                official_material_id,
                key["terrain_id"],
            )
            damages, lower_clamps, upper_clamps = interpolate_return_period_damage(
                [key[column] for column in GUST_COLUMNS],
                curve["wind_speed_mph"].to_numpy(),
                curve["building_loss_ratio"].to_numpy(),
            )
            expected = float(integrate_damage_matrix(damages.reshape(1, -1))[0])
            row = {
                **{column: key[column] for column in key_columns},
                "official_material_id": official_material_id,
                "fragility_proxy_id": proxy_id,
                "year_1_expected_damage_ratio": expected,
                "year_1_expected_damage_percent": expected * 100.0,
                "lower_clamp_count": lower_clamps,
                "upper_clamp_count": upper_clamps,
            }
            row.update(
                {
                    f"rp_{period}_damage_ratio": float(damage)
                    for period, damage in zip(RETURN_PERIODS, damages)
                }
            )
            calculation_rows.append(row)
    lookup = pl.DataFrame(calculation_rows).with_columns(
        pl.lit(EXPECTED_DAMAGE_VERSION).alias("dataset_version"),
        pl.lit(FRAGILITY_MAPPING_VERSION).alias("mapping_version"),
        pl.lit(True).alias("fragility_proxy_applied"),
        pl.lit(True).alias("tail_assumption"),
        pl.lit("equal_average_unresolved_variants_within_terrain").alias(
            "curve_aggregation_method"
        ),
    )
    _validate_lookup(lookup, unique_keys.select(key_columns))
    asset_results = _join_asset_results(assets_with_hazard, lookup)
    return lookup, asset_results


def build_asset_scoped_damage(
    asset_workbook_path: Path,
    wind_csv_path: Path,
    climate_zone_path: Path,
    fragility_root: Path,
    output_root: Path,
    sheet_name: str = "Sheet1",
) -> Dict[str, object]:
    """Run the repeatable asset-scoped pipeline and publish lookup/results/manifest."""
    assets = attach_climate_zones(
        attach_nearest_wind_grid(read_asset_workbook(asset_workbook_path, sheet_name), wind_csv_path),
        climate_zone_path,
    )
    fragility_identity = _fragility_identity(assets, fragility_root)
    calculation_cache_key = _cache_key(assets, fragility_identity)
    asset_run_key = _asset_run_key(assets)
    version_root = output_root / f"version={EXPECTED_DAMAGE_VERSION}"
    cache_root = version_root / "calculation_cache" / f"cache_key={calculation_cache_key}"
    run_root = version_root / "asset_runs" / f"asset_run_key={asset_run_key}"
    lookup_path = cache_root / "unique_location_damage_lookup.parquet"
    asset_results_path = run_root / "asset_material_damage_results.parquet"
    manifest_path = run_root / "build_manifest.json"
    if asset_results_path.is_file() and manifest_path.is_file():
        return json.loads(manifest_path.read_text(encoding="utf-8"))

    if lookup_path.is_file():
        lookup = pl.read_parquet(lookup_path)
        asset_results = _join_asset_results(assets, lookup)
        calculation_cache_hit = True
    else:
        lookup, asset_results = calculate_asset_damage(assets, fragility_root)
        cache_root.mkdir(parents=True, exist_ok=True)
        _atomic_write_parquet(lookup, lookup_path)
        calculation_cache_hit = False
    run_root.mkdir(parents=True, exist_ok=True)
    _atomic_write_parquet(asset_results, asset_results_path)
    manifest = {
        "dataset_version": EXPECTED_DAMAGE_VERSION,
        "mapping_version": FRAGILITY_MAPPING_VERSION,
        "fragility_identity": fragility_identity,
        "calculation_cache_key": calculation_cache_key,
        "calculation_cache_hit": calculation_cache_hit,
        "asset_run_key": asset_run_key,
        "asset_workbook": str(asset_workbook_path.resolve()),
        "asset_count": assets.height,
        "unique_calculation_key_count": assets.select(
            "wind_grid_id", "climate_zone", "lookup_roof_age", "terrain_id"
        ).unique().height,
        "material_result_count": asset_results.height,
        "materials": list(MATERIAL_PROXY_IDS),
        "terrain_labels": {str(key): value for key, value in TERRAIN_LABELS.items()},
        "lookup_path": str(lookup_path.resolve()),
        "asset_results_path": str(asset_results_path.resolve()),
        "wind_input": "nearest CONUS404 center; 3-second gust point estimates",
        "aep_integration": "piecewise trapezoid with endpoints (1,0) and (0,1)",
        "tail_assumption": True,
    }
    temporary_manifest = manifest_path.with_suffix(".json.tmp")
    temporary_manifest.write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    temporary_manifest.replace(manifest_path)
    return manifest


def _resolve_terrain(
    record: dict,
    label_column: Optional[str],
    id_column: Optional[str],
    row_number: int,
) -> int:
    label_value = record.get(label_column) if label_column else None
    id_value = record.get(id_column) if id_column else None
    terrain_from_label = None
    terrain_from_id = None
    if label_value is not None and str(label_value).strip():
        terrain_from_label = TERRAIN_IDS_BY_LABEL.get(str(label_value).strip().casefold())
        if terrain_from_label is None:
            raise ValueError(f"row {row_number}: unknown terrain label {label_value!r}")
    if id_value is not None and str(id_value).strip():
        number = _number(id_value, row_number, id_column or "terrain_id")
        if not number.is_integer() or int(number) not in TERRAIN_LABELS:
            raise ValueError(f"row {row_number}: terrain_id must be an integer from 1 to 5")
        terrain_from_id = int(number)
    if terrain_from_label is None and terrain_from_id is None:
        raise ValueError(f"row {row_number}: terrain is required")
    if terrain_from_label and terrain_from_id and terrain_from_label != terrain_from_id:
        raise ValueError(f"row {row_number}: terrain label and numeric ID do not agree")
    return terrain_from_label if terrain_from_label is not None else terrain_from_id  # type: ignore[return-value]


def _cache_key(assets: pl.DataFrame, fragility_identity: str = "unversioned") -> str:
    key_columns = [
        "wind_grid_id",
        "climate_zone",
        "lookup_roof_age",
        "terrain_id",
        *GUST_COLUMNS,
    ]
    payload = {
        "dataset_version": EXPECTED_DAMAGE_VERSION,
        "mapping_version": FRAGILITY_MAPPING_VERSION,
        "fragility_identity": fragility_identity,
        "keys": assets.select(key_columns).unique().sort(key_columns).to_dicts(),
    }
    return hashlib.sha256(
        json.dumps(payload, sort_keys=True, separators=(",", ":")).encode("utf-8")
    ).hexdigest()[:20]


def _fragility_identity(assets: pl.DataFrame, fragility_root: Path) -> str:
    selected_partitions = assets.select("climate_zone", "lookup_roof_age").unique().sort(
        "climate_zone", "lookup_roof_age"
    )
    files = []
    for row in selected_partitions.iter_rows(named=True):
        partition = fragility_root / row["climate_zone"] / f"age_{row['lookup_roof_age']:02d}"
        for filename in (
            "curves_hu.parquet",
            "curve_points_hu.parquet",
            "curve_attributes_hu.parquet",
        ):
            path = partition / filename
            if not path.is_file():
                raise FileNotFoundError(path)
            stat = path.stat()
            files.append(
                {
                    "path": str(path.resolve()),
                    "size": stat.st_size,
                    "mtime_ns": stat.st_mtime_ns,
                }
            )
    return hashlib.sha256(
        json.dumps(files, sort_keys=True, separators=(",", ":")).encode("utf-8")
    ).hexdigest()


def _asset_run_key(assets: pl.DataFrame) -> str:
    columns = [
        "asset_id",
        "latitude",
        "longitude",
        "current_roof_type",
        "input_roof_age",
        "terrain_id",
        "wind_grid_id",
        "climate_zone",
    ]
    payload = assets.select(columns).sort("asset_id").to_dicts()
    return hashlib.sha256(
        json.dumps(payload, sort_keys=True, separators=(",", ":")).encode("utf-8")
    ).hexdigest()[:20]


def _join_asset_results(assets: pl.DataFrame, lookup: pl.DataFrame) -> pl.DataFrame:
    key_columns = ["wind_grid_id", "climate_zone", "lookup_roof_age", "terrain_id"]
    required_keys = assets.select(key_columns).unique()
    _validate_lookup(lookup, required_keys)
    results = assets.join(
        lookup,
        on=key_columns,
        how="left",
        validate="m:m",
    )
    _validate_asset_results(assets, results)
    return results


def _validate_lookup(lookup: pl.DataFrame, required_keys: pl.DataFrame) -> None:
    key_columns = ["wind_grid_id", "climate_zone", "lookup_roof_age", "terrain_id"]
    expected_materials = set(MATERIAL_PROXY_IDS)
    missing_keys = required_keys.join(
        lookup.select(key_columns).unique(), on=key_columns, how="anti"
    )
    if missing_keys.height:
        raise ValueError(f"damage lookup is missing {missing_keys.height} required calculation keys")
    groups = lookup.group_by(key_columns).agg(
        pl.len().alias("row_count"),
        pl.col("official_material_id").n_unique().alias("material_count"),
        pl.col("official_material_id").alias("materials"),
    )
    invalid = groups.filter(
        (pl.col("row_count") != len(expected_materials))
        | (pl.col("material_count") != len(expected_materials))
    )
    if invalid.height or any(set(materials) != expected_materials for materials in groups["materials"]):
        raise ValueError("every calculation key must contain exactly Asphalt, Metal, and Tile")


def _validate_asset_results(assets: pl.DataFrame, results: pl.DataFrame) -> None:
    if results.height != assets.height * len(MATERIAL_PROXY_IDS):
        raise ValueError("asset damage join did not produce exactly three rows per asset")
    invalid = results.group_by("asset_id").agg(
        pl.len().alias("row_count"),
        pl.col("official_material_id").n_unique().alias("material_count"),
    ).filter(
        (pl.col("row_count") != len(MATERIAL_PROXY_IDS))
        | (pl.col("material_count") != len(MATERIAL_PROXY_IDS))
    )
    if invalid.height:
        raise ValueError("every asset must contain exactly three unique material results")


def _unit_sphere_xyz(latitude: np.ndarray, longitude: np.ndarray) -> np.ndarray:
    latitude_radians = np.radians(latitude)
    longitude_radians = np.radians(longitude)
    cosine_latitude = np.cos(latitude_radians)
    return np.column_stack(
        (
            cosine_latitude * np.cos(longitude_radians),
            cosine_latitude * np.sin(longitude_radians),
            np.sin(latitude_radians),
        )
    )


def _number(value: object, row_number: int, column: str) -> float:
    try:
        number = float(value)  # type: ignore[arg-type]
    except (TypeError, ValueError) as error:
        raise ValueError(f"row {row_number}: {column} must be numeric") from error
    if not np.isfinite(number):
        raise ValueError(f"row {row_number}: {column} must be finite")
    return number


def _first_present(headers: Iterable[str], choices: Sequence[str]) -> Optional[str]:
    header_set = set(headers)
    return next((choice for choice in choices if choice in header_set), None)


def _atomic_write_parquet(frame: pl.DataFrame, path: Path) -> None:
    temporary_path = path.with_suffix(".parquet.tmp")
    frame.write_parquet(temporary_path, compression="zstd", statistics=True)
    temporary_path.replace(path)


def _parse_args(arguments: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--assets", type=Path, required=True)
    parser.add_argument("--sheet", default="Sheet1")
    parser.add_argument("--wind-csv", type=Path, required=True)
    parser.add_argument("--climate-zones", type=Path, required=True)
    parser.add_argument("--fragility-root", type=Path, required=True)
    parser.add_argument("--output-root", type=Path, required=True)
    return parser.parse_args(arguments)


def main() -> None:
    args = _parse_args()
    manifest = build_asset_scoped_damage(
        args.assets,
        args.wind_csv,
        args.climate_zones,
        args.fragility_root,
        args.output_root,
        sheet_name=args.sheet,
    )
    print(json.dumps(manifest, indent=2))


if __name__ == "__main__":
    main()